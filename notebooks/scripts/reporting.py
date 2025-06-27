"""
05_reporting.py
Report Generation Script for Nepal Earthquake Damage Assessment
Creates comprehensive PDF reports, Excel summaries, and GIS-ready outputs
"""

import pandas as pd
import numpy as np
import geopandas as gpd
from pathlib import Path
import json
from datetime import datetime
import matplotlib.pyplot as plt
import seaborn as sns
from typing import Dict, Optional, List
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
import rasterio
from rasterio.crs import CRS
import warnings
warnings.filterwarnings('ignore')

import logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class ReportGenerator:
    """
    Generate comprehensive reports for earthquake damage assessment
    """
    
    def __init__(self, config_path: str = 'config.json'):
        """Initialize report generator"""
        self.config = self._load_config(config_path)
        self.data_dir = Path(self.config['data_dir'])
        self.results_dir = self.data_dir / 'results'
        self.reports_dir = self.results_dir / 'reports'
        self.reports_dir.mkdir(exist_ok=True)
        
        # Report metadata
        self.metadata = {
            'title': 'Nepal Earthquake Damage Assessment Report',
            'organization': 'Disaster Response Team',
            'date': datetime.now().strftime('%Y-%m-%d'),
            'version': '1.0'
        }
    
    def _load_config(self, config_path: str) -> Dict:
        """Load configuration"""
        with open(config_path, 'r') as f:
            return json.load(f)
    
    def generate_pdf_report(self, output_path: Optional[str] = None) -> str:
        """Generate comprehensive PDF report"""
        logger.info("Generating PDF report")
        
        if not output_path:
            output_path = self.reports_dir / f"damage_assessment_report_{datetime.now().strftime('%Y%m%d')}.pdf"
        
        # Create PDF document
        doc = SimpleDocTemplate(str(output_path), pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#34495e'),
            spaceAfter=12
        )
        
        # Title page
        story.append(Paragraph(self.metadata['title'], title_style))
        story.append(Spacer(1, 0.5*inch))
        
        # Event information
        event_info = f"""
        <para align="center">
        <b>Event Date:</b> {self.config['earthquake']['date']}<br/>
        <b>Magnitude:</b> {self.config['earthquake']['magnitude']}<br/>
        <b>Epicenter:</b> {self.config['earthquake']['epicenter'][1]:.4f}°N, {self.config['earthquake']['epicenter'][0]:.4f}°E<br/>
        <b>Affected Districts:</b> {', '.join(self.config['earthquake']['affected_districts'])}<br/>
        <b>Report Date:</b> {self.metadata['date']}
        </para>
        """
        story.append(Paragraph(event_info, styles['Normal']))
        story.append(PageBreak())
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", heading_style))
        
        # Load statistics
        stats_path = self.results_dir / 'damage_assessment_report.json'
        if stats_path.exists():
            with open(stats_path, 'r') as f:
                stats = json.load(f)
            
            summary_text = self._generate_executive_summary(stats)
            story.append(Paragraph(summary_text, styles['BodyText']))
            story.append(Spacer(1, 0.3*inch))
        
        # Damage Statistics Table
        story.append(Paragraph("Damage Statistics", heading_style))
        
        if 'damage_distribution' in stats:
            damage_table_data = [['Damage Level', 'Area (km²)', 'Percentage']]
            
            for level, data in stats['damage_distribution'].items():
                damage_table_data.append([
                    level,
                    f"{data['area_km2']:.2f}",
                    f"{data['percentage']:.1f}%"
                ])
            
            damage_table = Table(damage_table_data)
            damage_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(damage_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Add visualizations
        viz_dir = self.results_dir / 'visualizations'
        
        # Before/After comparison
        if (viz_dir / 'before_after_comparison.png').exists():
            story.append(Paragraph("Before/After Comparison", heading_style))
            img = Image(str(viz_dir / 'before_after_comparison.png'), 
                       width=6*inch, height=3*inch)
            story.append(img)
            story.append(PageBreak())
        
        # Damage map
        if (viz_dir / 'damage_map.png').exists():
            story.append(Paragraph("Damage Assessment Map", heading_style))
            img = Image(str(viz_dir / 'damage_map.png'), 
                       width=5*inch, height=4*inch)
            story.append(img)
            story.append(Spacer(1, 0.3*inch))
        
        # Building damage summary
        if 'building_damage' in stats:
            story.append(Paragraph("Building Damage Assessment", heading_style))
            
            building_data = [['Damage Category', 'Number of Buildings']]
            for category, count in stats['building_damage']['damage_summary'].items():
                building_data.append([category, str(count)])
            
            building_table = Table(building_data)
            building_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(building_table)
            story.append(PageBreak())
        
        # Methodology section
        story.append(Paragraph("Methodology", heading_style))
        methodology_text = """
        This damage assessment was conducted using state-of-the-art satellite imagery analysis 
        and GeoAI techniques. The methodology included:
        
        • Multi-temporal satellite imagery analysis (Sentinel-2, Landsat 8/9)
        • Spectral change detection (NDVI, NBR, NDBI indices)
        • Texture analysis for structural damage assessment
        • Machine learning classification using Random Forest and XGBoost
        • SAR coherence analysis for all-weather monitoring
        • Building-level damage assessment using high-resolution imagery
        • Landslide detection using topographic and spectral analysis
        """
        story.append(Paragraph(methodology_text, styles['BodyText']))
        
        # Data sources
        story.append(Paragraph("Data Sources", heading_style))
        data_sources_text = """
        • Sentinel-2 Level-2A (ESA Copernicus Programme)
        • Sentinel-1 SAR GRD (ESA Copernicus Programme)
        • Landsat 8/9 Collection 2 (USGS)
        • OpenStreetMap building footprints
        • SRTM Digital Elevation Model (NASA)
        • Population density (CIESIN GPW v4.11)
        """
        story.append(Paragraph(data_sources_text, styles['BodyText']))
        
        # Build PDF
        doc.build(story)
        logger.info(f"PDF report saved to {output_path}")
        
        return str(output_path)
    
    def generate_excel_report(self, output_path: Optional[str] = None) -> str:
        """Generate detailed Excel report with multiple sheets"""
        logger.info("Generating Excel report")
        
        if not output_path:
            output_path = self.reports_dir / f"damage_assessment_data_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        # Create workbook
        wb = Workbook()
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 1. Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Add title
        ws_summary.merge_cells('A1:E1')
        ws_summary['A1'] = "Nepal Earthquake Damage Assessment Summary"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary['A1'].alignment = Alignment(horizontal="center")
        
        # Event information
        event_data = [
            ['Event Information', ''],
            ['Date', self.config['earthquake']['date']],
            ['Magnitude', self.config['earthquake']['magnitude']],
            ['Epicenter Lat', self.config['earthquake']['epicenter'][1]],
            ['Epicenter Lon', self.config['earthquake']['epicenter'][0]],
            ['Affected Districts', ', '.join(self.config['earthquake']['affected_districts'])],
            ['', ''],
            ['Assessment Information', ''],
            ['Assessment Date', datetime.now().strftime('%Y-%m-%d')],
            ['Data Sources', 'Sentinel-2, Landsat 8/9, Sentinel-1']
        ]
        
        for row_idx, row_data in enumerate(event_data, start=3):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 1:
                    cell.font = Font(bold=True)
        
        # 2. Damage statistics sheet
        ws_damage = wb.create_sheet("Damage Statistics")
        
        # Load statistics
        stats_path = self.results_dir / 'damage_assessment_report.json'
        if stats_path.exists():
            with open(stats_path, 'r') as f:
                stats = json.load(f)
            
            if 'damage_distribution' in stats:
                # Headers
                headers = ['Damage Level', 'Area (km²)', 'Pixels', 'Percentage']
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws_damage.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Data
                row_idx = 2
                for level, data in stats['damage_distribution'].items():
                    ws_damage.cell(row=row_idx, column=1, value=level).border = border
                    ws_damage.cell(row=row_idx, column=2, value=data['area_km2']).border = border
                    ws_damage.cell(row=row_idx, column=3, value=data['pixels']).border = border
                    ws_damage.cell(row=row_idx, column=4, value=data['percentage']).border = border
                    row_idx += 1
                
                # Add chart
                chart = PieChart()
                chart.title = "Damage Distribution"
                chart.add_data(Reference(ws_damage, min_col=4, min_row=2, max_row=row_idx-1))
                chart.set_categories(Reference(ws_damage, min_col=1, min_row=2, max_row=row_idx-1))
                ws_damage.add_chart(chart, "F2")
        
        # 3. Building damage sheet
        building_damage_path = self.results_dir / 'building_damage_assessment.geojson'
        if building_damage_path.exists():
            ws_buildings = wb.create_sheet("Building Damage")
            
            # Load building data
            buildings_gdf = gpd.read_file(building_damage_path)
            buildings_df = pd.DataFrame(buildings_gdf.drop(columns='geometry'))
            
            # Write headers
            for col_idx, column in enumerate(buildings_df.columns, start=1):
                cell = ws_buildings.cell(row=1, column=col_idx, value=column)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # Write data
            for row_idx, row in enumerate(buildings_df.itertuples(index=False), start=2):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws_buildings.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
            
            # Add summary
            summary_row = len(buildings_df) + 4
            ws_buildings.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True)
            
            damage_summary = buildings_df['damage_class'].value_counts()
            for idx, (damage_class, count) in enumerate(damage_summary.items()):
                ws_buildings.cell(row=summary_row + idx + 1, column=1, value=damage_class)
                ws_buildings.cell(row=summary_row + idx + 1, column=2, value=count)
        
        # 4. District summary sheet
        ws_districts = wb.create_sheet("District Summary")
        
        # Create district summary data
        district_data = []
        for district in self.config['earthquake']['affected_districts']:
            district_data.append({
                'District': district,
                'Status': 'Severely Affected' if district == 'Jajarkot' else 'Affected',
                'Priority': 'High' if district in ['Jajarkot', 'Rukum West'] else 'Medium'
            })
        
        district_df = pd.DataFrame(district_data)
        
        # Write to sheet
        for col_idx, column in enumerate(district_df.columns, start=1):
            cell = ws_districts.cell(row=1, column=col_idx, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        for row_idx, row in enumerate(district_df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws_districts.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
        
        # 5. Recommendations sheet
        ws_recommendations = wb.create_sheet("Recommendations")
        
        recommendations = self._generate_recommendations(stats if 'stats' in locals() else {})
        
        ws_recommendations.merge_cells('A1:D1')
        ws_recommendations['A1'] = "Recommendations for Response and Recovery"
        ws_recommendations['A1'].font = Font(bold=True, size=14)
        
        for idx, rec in enumerate(recommendations, start=3):
            ws_recommendations.cell(row=idx, column=1, value=f"{idx-2}.")
            ws_recommendations.merge_cells(f'B{idx}:D{idx}')
            ws_recommendations.cell(row=idx, column=2, value=rec)
        
        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel report saved to {output_path}")
        
        return str(output_path)
    
    def generate_gis_outputs(self, output_dir: Optional[str] = None) -> Dict[str, str]:
        """Generate GIS-ready outputs (shapefiles, GeoPackage, KML)"""
        logger.info("Generating GIS outputs")
        
        if not output_dir:
            output_dir = self.reports_dir / 'gis_outputs'
        output_dir = Path(output_dir)
        output_dir.mkdir(exist_ok=True)
        
        outputs = {}
        
        # 1. Convert damage raster to vector
        damage_raster = self.results_dir / 'damage_classification.tif'
        if damage_raster.exists():
            damage_vector = self._rasterize_damage(damage_raster)
            
            # Save as different formats
            # Shapefile
            shp_path = output_dir / 'damage_assessment.shp'
            damage_vector.to_file(shp_path)
            outputs['shapefile'] = str(shp_path)
            
            # GeoPackage
            gpkg_path = output_dir / 'damage_assessment.gpkg'
            damage_vector.to_file(gpkg_path, driver='GPKG')
            outputs['geopackage'] = str(gpkg_path)
            
            # KML for Google Earth
            kml_path = output_dir / 'damage_assessment.kml'
            damage_vector_wgs84 = damage_vector.to_crs('EPSG:4326')
            damage_vector_wgs84.to_file(kml_path, driver='KML')
            outputs['kml'] = str(kml_path)
        
        # 2. Building damage outputs
        building_damage = self.results_dir / 'building_damage_assessment.geojson'
        if building_damage.exists():
            buildings_gdf = gpd.read_file(building_damage)
            
            # Shapefile
            buildings_shp = output_dir / 'building_damage.shp'
            buildings_gdf.to_file(buildings_shp)
            outputs['buildings_shapefile'] = str(buildings_shp)
            
            # CSV with coordinates
            buildings_csv = output_dir / 'building_damage.csv'
            buildings_df = buildings_gdf.copy()
            buildings_df['longitude'] = buildings_df.geometry.x
            buildings_df['latitude'] = buildings_df.geometry.y
            buildings_df.drop(columns='geometry').to_csv(buildings_csv, index=False)
            outputs['buildings_csv'] = str(buildings_csv)
        
        # 3. Create metadata file
        metadata_path = output_dir / 'metadata.txt'
        self._create_gis_metadata(metadata_path, outputs)
        outputs['metadata'] = str(metadata_path)
        
        logger.info(f"GIS outputs saved to {output_dir}")
        return outputs
    
    def generate_web_report(self, output_path: Optional[str] = None) -> str:
        """Generate web-based report package"""
        logger.info("Generating web report package")
        
        if not output_path:
            output_path = self.reports_dir / 'web_report'
        output_path = Path(output_path)
        output_path.mkdir(exist_ok=True)
        
        # Copy visualizations
        viz_dir = self.results_dir / 'visualizations'
        if viz_dir.exists():
            import shutil
            for file in viz_dir.glob('*'):
                shutil.copy2(file, output_path / file.name)
        
        # Create data.json with all statistics
        stats_path = self.results_dir / 'damage_assessment_report.json'
        if stats_path.exists():
            with open(stats_path, 'r') as f:
                stats = json.load(f)
            
            # Add additional metadata
            stats['metadata'] = self.metadata
            stats['earthquake_info'] = self.config['earthquake']
            
            with open(output_path / 'data.json', 'w') as f:
                json.dump(stats, f, indent=2)
        
        # Create index.html
        index_content = """
        <!DOCTYPE html>
        <html>
        <head>
            <title>Nepal Earthquake Damage Assessment</title>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <link rel="stylesheet" href="https://unpkg.com/leaflet@1.7.1/dist/leaflet.css">
            <script src="https://unpkg.com/leaflet@1.7.1/dist/leaflet.js"></script>
            <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
            <style>
                body { font-family: Arial, sans-serif; margin: 0; padding: 0; }
                #map { height: 600px; width: 100%; }
                .container { max-width: 1200px; margin: 0 auto; padding: 20px; }
                .stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 20px 0; }
                .stat-card { background: #f0f0f0; padding: 20px; border-radius: 8px; text-align: center; }
                .stat-value { font-size: 2em; font-weight: bold; color: #333; }
                .stat-label { color: #666; margin-top: 5px; }
            </style>
        </head>
        <body>
            <div class="container">
                <h1>Nepal Earthquake Damage Assessment</h1>
                <div id="map"></div>
                <div class="stats-grid" id="stats"></div>
                <div id="charts"></div>
            </div>
            
            <script>
                // Load and display data
                fetch('data.json')
                    .then(response => response.json())
                    .then(data => {
                        // Initialize map
                        var map = L.map('map').setView(
                            [data.earthquake_info.epicenter[1], data.earthquake_info.epicenter[0]], 
                            9
                        );
                        
                        L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png').addTo(map);
                        
                        // Add epicenter marker
                        L.marker([data.earthquake_info.epicenter[1], data.earthquake_info.epicenter[0]])
                            .addTo(map)
                            .bindPopup('Epicenter<br>Magnitude: ' + data.earthquake_info.magnitude);
                        
                        // Display statistics
                        displayStats(data);
                    });
                
                function displayStats(data) {
                    const statsContainer = document.getElementById('stats');
                    
                    if (data.damage_distribution) {
                        Object.entries(data.damage_distribution).forEach(([level, info]) => {
                            const card = document.createElement('div');
                            card.className = 'stat-card';
                            card.innerHTML = `
                                <div class="stat-value">${info.area_km2.toFixed(1)} km²</div>
                                <div class="stat-label">${level}</div>
                            `;
                            statsContainer.appendChild(card);
                        });
                    }
                }
            </script>
        </body>
        </html>
        """
        
        with open(output_path / 'index.html', 'w') as f:
            f.write(index_content)
        
        logger.info(f"Web report saved to {output_path}")
        return str(output_path)
    
    def _generate_executive_summary(self, stats: Dict) -> str:
        """Generate executive summary text"""
        total_area = stats.get('total_area_km2', 0)
        
        summary = f"""
        A comprehensive damage assessment was conducted following the magnitude {self.config['earthquake']['magnitude']} 
        earthquake that struck {self.config['earthquake']['affected_districts'][0]} District, Nepal on 
        {self.config['earthquake']['date']}. The assessment utilized multi-temporal satellite imagery analysis 
        and advanced GeoAI techniques to evaluate the extent and severity of damage across the affected region.
        
        Key findings include:
        """
        
        if 'damage_distribution' in stats:
            severe_damage = stats['damage_distribution'].get('Severe', {}).get('area_km2', 0)
            high_damage = stats['damage_distribution'].get('High', {}).get('area_km2', 0)
            total_significant = severe_damage + high_damage
            
            summary += f"""
            • Total assessed area: {total_area:.2f} km²
            • Severely damaged area: {severe_damage:.2f} km² 
            • Total significant damage (High + Severe): {total_significant:.2f} km²
            """
        
        if 'building_damage' in stats:
            total_buildings = stats['building_damage']['total_buildings']
            summary += f"""
            • Buildings assessed: {total_buildings}
            • Buildings with major damage or destroyed: {
                stats['building_damage']['damage_summary'].get('Major', 0) + 
                stats['building_damage']['damage_summary'].get('Destroyed', 0)
            }
            """
        
        if 'landslide_area_km2' in stats:
            summary += f"""
            • Landslide-affected area: {stats['landslide_area_km2']:.2f} km²
            """
        
        summary += """
        
        The assessment reveals significant structural damage concentrated in the epicentral region, 
        with widespread impacts on residential buildings, infrastructure, and natural terrain stability. 
        Immediate humanitarian assistance and structural safety assessments are recommended for the 
        most severely affected areas.
        """
        
        return summary
    
    def _generate_recommendations(self, stats: Dict) -> List[str]:
        """Generate recommendations based on assessment results"""
        recommendations = [
            "Immediate Actions:",
            "• Deploy search and rescue teams to areas identified with severe building damage",
            "• Establish temporary shelters for displaced populations in affected districts",
            "• Conduct detailed structural engineering assessments of damaged buildings",
            "• Monitor identified landslide areas for continued instability",
            "",
            "Short-term Recovery:",
            "• Develop debris removal and management plan for destroyed structures",
            "• Implement temporary housing solutions for affected families",
            "• Restore critical infrastructure including roads, water, and electricity",
            "• Provide psychosocial support services to affected communities",
            "",
            "Long-term Reconstruction:",
            "• Enforce seismic building codes in reconstruction efforts",
            "• Develop landslide hazard maps and implement slope stabilization",
            "• Create community-based disaster preparedness programs",
            "• Establish early warning systems for future seismic events"
        ]
        
        # Add specific recommendations based on statistics
        if 'damage_distribution' in stats:
            severe_pct = stats['damage_distribution'].get('Severe', {}).get('percentage', 0)
            if severe_pct > 10:
                recommendations.insert(2, 
                    "• URGENT: Over 10% of assessed area shows severe damage - prioritize emergency response")
        
        return recommendations
    
    def _rasterize_damage(self, raster_path: str) -> gpd.GeoDataFrame:
        """Convert damage raster to vector polygons"""
        from rasterio.features import shapes
        
        with rasterio.open(raster_path) as src:
            image = src.read(1)
            transform = src.transform
            
            # Get shapes
            results = []
            for geom, value in shapes(image.astype(np.int16), transform=transform):
                if value > 0:  # Only damaged areas
                    results.append({
                        'geometry': geom,
                        'damage_level': int(value),
                        'damage_class': ['No damage', 'Low', 'Moderate', 'High', 'Severe'][int(value)]
                    })
            
            gdf = gpd.GeoDataFrame(results)
            gdf.crs = src.crs
            
        return gdf
    
    def _create_gis_metadata(self, output_path: str, outputs: Dict):
        """Create metadata file for GIS outputs"""
        metadata_content = f"""
        NEPAL EARTHQUAKE DAMAGE ASSESSMENT - GIS DATA METADATA
        =====================================================
        
        Creation Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        
        Event Information:
        - Date: {self.config['earthquake']['date']}
        - Magnitude: {self.config['earthquake']['magnitude']}
        - Epicenter: {self.config['earthquake']['epicenter']}
        
        Coordinate Reference System:
        - Original data: WGS 84 / UTM Zone 44N (EPSG:32644)
        - KML files: WGS 84 (EPSG:4326)
        
        Files Generated:
        """
        
        for key, path in outputs.items():
            if key != 'metadata':
                metadata_content += f"\n- {key}: {Path(path).name}"
        
        metadata_content += """
        
        Attribute Descriptions:
        - damage_level: Numeric damage classification (0-4)
        - damage_class: Text damage classification (No damage, Low, Moderate, High, Severe)
        - damage_score: Continuous damage score (0-1) for buildings
        - building_id: Unique identifier for each building
        
        Data Sources:
        - Sentinel-2 imagery (ESA Copernicus)
        - Landsat 8/9 imagery (USGS)
        - OpenStreetMap building footprints
        
        Processing Methodology:
        - Multi-temporal change detection
        - Machine learning classification
        - GeoAI-based damage assessment
        
        Contact: {organization}
        """
        
        with open(output_path, 'w') as f:
            f.write(metadata_content.format(
                organization=self.metadata.get('organization', 'Disaster Response Team')
            ))
    
    def generate_all_reports(self):
        """Generate all report types"""
        logger.info("Generating all reports")
        
        reports = {}
        
        # PDF report
        try:
            reports['pdf'] = self.generate_pdf_report()
        except Exception as e:
            logger.error(f"Failed to generate PDF report: {e}")
        
        # Excel report
        try:
            reports['excel'] = self.generate_excel_report()
        except Exception as e:
            logger.error(f"Failed to generate Excel report: {e}")
        
        # GIS outputs
        try:
            reports['gis'] = self.generate_gis_outputs()
        except Exception as e:
            logger.error(f"Failed to generate GIS outputs: {e}")
        
        # Web report
        try:
            reports['web'] = self.generate_web_report()
        except Exception as e:
            logger.error(f"Failed to generate web report: {e}")
        
        # Save report paths
        report_summary = self.reports_dir / 'generated_reports.json'
        with open(report_summary, 'w') as f:
            json.dump(reports, f, indent=2)
        
        logger.info(f"All reports generated. Summary saved to {report_summary}")
        return reports


def main():
    """Main execution function"""
    # Initialize report generator
    generator = ReportGenerator('config.json')
    
    # Generate all reports
    reports = generator.generate_all_reports()
    
    logger.info("Report generation complete!")
    logger.info("Generated reports:")
    for report_type, path in reports.items():
        if isinstance(path, dict):
            logger.info(f"  {report_type}:")
            for sub_type, sub_path in path.items():
                logger.info(f"    - {sub_type}: {sub_path}")
        else:
            logger.info(f"  {report_type}: {path}")


if __name__ == "__main__":
    main()